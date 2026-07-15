import streamlit as st
import os, re, requests
from pypdf import PdfReader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from sentence_transformers import SentenceTransformer
import chromadb
from rank_bm25 import BM25Okapi

# ── Settings ──────────────────────────────────────────────────────────────────
COLLECTION_NAME = "documentqa_ui"
CHROMA_PATH = "./vectorstore_ui"
OLLAMA_MODEL = "llama3.2"
TOP_K = 3
EXPLETIVES = ["damn", "shit", "fuck", "ass", "crap", "bastard", "bitch", "hell"]
RELEVANCE_THRESHOLD = 1.0
# ──────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="DocumentQA", layout="wide")
st.title("DocumentQA")
st.caption("Upload a document and ask questions about its contents.")

@st.cache_resource
def load_embedding_model():
    return SentenceTransformer("all-MiniLM-L6-v2")

embedding_model = load_embedding_model()

def check_expletives(text):
    return any(word in text.lower() for word in EXPLETIVES)

def extract_text(file, ext):
    if ext == ".pdf":
        reader = PdfReader(file)
        full_text = ""
        for page_num, page in enumerate(reader.pages, start=1):
            text = page.extract_text()
            if text:
                full_text += f"\n[Page {page_num}]\n{text}"
        return full_text
    elif ext in (".txt", ".md"):
        return file.read().decode("utf-8")
    elif ext == ".docx":
        import docx
        doc = docx.Document(file)
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
    elif ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(file)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " | ".join([str(c) if c else "" for c in row]) + "\n"
        return text
    return ""

def detect_section(chunk):
    patterns = [
        r'\b(\d+\.\d+(?:\.\d+)*)\b',
        r'[Ss]ection\s+(\d+(?:\.\d+)*)',
        r'[Cc]lause\s+(\d+(?:\.\d+)*)',
    ]
    for pattern in patterns:
        match = re.search(pattern, chunk)
        if match:
            return match.group(0).strip()
    return None

def chunk_and_embed(text):
    splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
    chunks = splitter.split_text(text)
    embeddings = embedding_model.encode(chunks, show_progress_bar=False)
    os.makedirs(CHROMA_PATH, exist_ok=True)
    client = chromadb.PersistentClient(path=CHROMA_PATH)
    try:
        client.delete_collection(COLLECTION_NAME)
    except:
        pass
    collection = client.create_collection(COLLECTION_NAME)
    collection.add(
        documents=chunks,
        embeddings=embeddings.tolist(),
        ids=[f"chunk_{i}" for i in range(len(chunks))],
        metadatas=[{"chunk_index": i} for i in range(len(chunks))]
    )

    # Build BM25 index for keyword search
    tokenized = [chunk.lower().split() for chunk in chunks]
    bm25 = BM25Okapi(tokenized)

    return collection, len(chunks), chunks, bm25

def vector_retrieve(query, collection):
    """Traditional RAG — semantic vector search"""
    query_embedding = embedding_model.encode([query]).tolist()
    results = collection.query(
        query_embeddings=query_embedding,
        n_results=TOP_K,
        include=["documents", "distances"]
    )
    return results["documents"][0], results["distances"][0]

def keyword_retrieve(query, chunks, bm25):
    """Vectorless RAG — BM25 keyword search"""
    tokenized_query = query.lower().split()
    scores = bm25.get_scores(tokenized_query)
    top_indices = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)[:TOP_K]
    return [chunks[i] for i in top_indices]

def hybrid_retrieve(query, collection, chunks, bm25):
    """Hybrid RAG — combine vector + keyword results"""
    vector_docs, distances = vector_retrieve(query, collection)
    keyword_docs = keyword_retrieve(query, chunks, bm25)

    # Combine and deduplicate
    seen = set()
    combined = []
    for doc in vector_docs + keyword_docs:
        key = doc[:100]  # use first 100 chars as key
        if key not in seen:
            seen.add(key)
            combined.append(doc)

    return combined[:TOP_K + 2], distances  # return top results

def is_relevant(distances):
    return min(distances) < RELEVANCE_THRESHOLD

def ask_ollama(prompt):
    try:
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False},
            timeout=120
        )
        return response.json()["response"]
    except:
        return "Error: Could not connect to Ollama. Make sure Ollama is running."

def build_prompt(query, docs):
    context = "\n\n".join([f"Chunk {i+1}:\n{c}" for i, c in enumerate(docs)])
    return f"""You are a helpful document analysis assistant. Answer the question below using only the context provided.

Instructions:
- Give a detailed, thorough answer. Do not be brief.
- If the answer contains multiple points, list them clearly using bullet points or numbered lists.
- If there are definitions, procedures, or rules, present them in a structured way.
- Always mention the page number and section number if they appear in the context.
- End your answer with a one-line summary of the key takeaway.

Context:
{context}

Question: {query}

Detailed Answer:"""

def build_summary_prompt(all_chunks):
    context = "\n\n".join([f"Chunk {i+1}:\n{c}" for i, c in enumerate(all_chunks[:8])])
    return f"""You are a helpful assistant. Based on the document excerpts below, write a clear and concise summary.
Cover the main topics, key points, and purpose of the document in 4-6 sentences.

Document excerpts:
{context}

Summary:"""

def get_references(docs):
    pages, sections = [], []
    for doc in docs:
        if isinstance(doc, str):
            pages.extend(re.findall(r'\[Page (\d+)\]', doc))
            section = detect_section(doc)
            if section:
                sections.append(section)
    page_str = "Pages: " + ", ".join(sorted(set(pages), key=int)) if pages else ""
    section_str = "Sections: " + ", ".join(sorted(set(sections))) if sections else ""
    return " | ".join(filter(None, [page_str, section_str]))

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### Upload Document")
    uploaded_file = st.file_uploader(
        "Supported: PDF, Word, Excel, TXT, Markdown",
        type=["pdf", "txt", "md", "docx", "xlsx"]
    )

    # Retrieval mode selector
    retrieval_mode = st.radio(
        "Retrieval Mode",
        ["Hybrid (Vector + Keyword)", "Vector only", "Keyword only"],
        index=0
    )

    if uploaded_file:
        ext = os.path.splitext(uploaded_file.name)[1].lower()
        st.write(f"File: **{uploaded_file.name}**")

        if st.button("Process Document"):
            with st.spinner("Processing..."):
                text = extract_text(uploaded_file, ext)
                if not text.strip():
                    st.error("Could not extract text from this file.")
                else:
                    collection, num_chunks, all_chunks, bm25 = chunk_and_embed(text)
                    st.session_state["collection"] = collection
                    st.session_state["doc_name"] = uploaded_file.name
                    st.session_state["num_chunks"] = num_chunks
                    st.session_state["all_chunks"] = all_chunks
                    st.session_state["bm25"] = bm25
                    st.session_state["messages"] = []
                    st.success(f"Done. {num_chunks} chunks indexed.")

    if "doc_name" in st.session_state:
        st.markdown("---")
        st.write(f"Active: **{st.session_state['doc_name']}**")
        st.caption(f"{st.session_state['num_chunks']} chunks in index")
        st.markdown("---")

        if st.button("Summarize Document"):
            with st.spinner("Generating summary..."):
                prompt = build_summary_prompt(st.session_state["all_chunks"])
                summary = ask_ollama(prompt)
                summary_msg = f"**Document Summary**\n\n{summary}"
                st.session_state["messages"].append({"role": "assistant", "content": summary_msg})
                st.rerun()

# ── Chat ──────────────────────────────────────────────────────────────────────
if "messages" not in st.session_state:
    st.session_state["messages"] = []

if "collection" not in st.session_state:
    st.info("Upload a document from the sidebar to get started.")
else:
    for msg in st.session_state["messages"]:
        with st.chat_message(msg["role"]):
            st.markdown(msg["content"])

    if prompt := st.chat_input("Ask a question about your document..."):
        with st.chat_message("user"):
            st.markdown(prompt)
        st.session_state["messages"].append({"role": "user", "content": prompt})

        if check_expletives(prompt):
            response = "Expletives are not allowed. Please rephrase your question politely."
        else:
            with st.spinner("Searching and generating answer..."):

                # Choose retrieval mode
                if retrieval_mode == "Hybrid (Vector + Keyword)":
                    docs, distances = hybrid_retrieve(
                        prompt,
                        st.session_state["collection"],
                        st.session_state["all_chunks"],
                        st.session_state["bm25"]
                    )
                elif retrieval_mode == "Vector only":
                    docs, distances = vector_retrieve(prompt, st.session_state["collection"])
                else:
                    docs = keyword_retrieve(prompt, st.session_state["all_chunks"], st.session_state["bm25"])
                    distances = [0.5]  # dummy distance for keyword-only

                if not is_relevant(distances):
                    response = "I can only answer questions related to the uploaded document. Your question does not appear to be relevant to its contents."
                else:
                    answer = ask_ollama(build_prompt(prompt, docs))
                    refs = get_references(docs)
                    response = answer
                    if refs:
                        response += f"\n\n**References:** {refs}"
                    response += f"\n\n*Retrieval mode: {retrieval_mode}*"

        with st.chat_message("assistant"):
            st.markdown(response)
        st.session_state["messages"].append({"role": "assistant", "content": response})